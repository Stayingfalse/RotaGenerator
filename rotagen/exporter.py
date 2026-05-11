from __future__ import annotations

import hashlib
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rotagen.models import DAYS, ScheduleResult, slot_label


# Pastel background colors for day groups (cycles if more than 6 days)
_DAY_COLORS = ["D6E4F7", "D6F7E4", "F7F0D6", "F7D6D6", "E4D6F7", "D6F7F3"]

# Maximum column width (characters) in the main rota sheet.
# Person-ID columns are intentionally narrower than the old comma-joined list
# columns, so 20 characters is sufficient and keeps the sheet compact.
_MAX_COL_WIDTH = 20


def _person_bg_fg(person_id: str) -> tuple[str, str]:
    """Return a consistent pastel (bg_hex, fg_hex) pair for a person_id.

    The background is derived from the MD5 hash of the ID, clamped to a
    mid-range so it is always readable.  The foreground (black or white) is
    chosen for maximum WCAG contrast against that background.
    """
    digest = hashlib.sha256(person_id.encode()).hexdigest()
    r = (int(digest[0:2], 16) % 128) + 100
    g = (int(digest[2:4], 16) % 128) + 100
    b = (int(digest[4:6], 16) % 128) + 100
    bg = f"{r:02X}{g:02X}{b:02X}"

    def _lin(c: int) -> float:
        v = c / 255.0
        return v / 12.92 if v <= 0.04045 else ((v + 0.055) / 1.055) ** 2.4

    lum = 0.2126 * _lin(r) + 0.7152 * _lin(g) + 0.0722 * _lin(b)
    fg = "000000" if lum > 0.179 else "FFFFFF"
    return bg, fg


def export_single_worksheet(
    result: ScheduleResult,
    queue_order: list[str] | None = None,
) -> bytes:
    """Export *result* to an XLSX workbook and return the raw bytes.

    Parameters
    ----------
    result:
        The schedule result to export.
    queue_order:
        Ordered list of queue names reflecting the configured priority.  Any
        queues present in *result* but absent from this list are appended
        alphabetically at the end.  Queues in *queue_order* that do not appear
        in *result* are silently dropped.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rota"

    # ---- Determine ordered queue list ----
    # Start from caller-supplied priority order, then append unseen queues.
    seen_in_result: set[str] = (
        {a.queue for a in result.assignments} | {c.queue for c in result.conflicts}
    )
    base_order: list[str] = list(queue_order) if queue_order else []
    for q in sorted(seen_in_result - set(base_order)):
        base_order.append(q)
    queues = [q for q in base_order if q in seen_in_result]

    # ---- Build assignment lookup ----
    # (day, slot, queue) -> sorted list of person_ids
    grid: dict[tuple[str, int, str], list[str]] = {}
    for a in result.assignments:
        key = (a.day, a.slot, a.queue)
        grid.setdefault(key, []).append(a.person_id)

    # ---- Days present (assignments OR conflicts), preserving canonical order ----
    days_with_data: set[str] = (
        {a.day for a in result.assignments} | {c.day for c in result.conflicts}
    )
    days_present = [d for d in DAYS if d in days_with_data]

    # ---- Column-count per queue ----
    # Each queue expands to N columns so that every required person gets their
    # own cell.  N = max over all (day, slot) of: people assigned OR total
    # required (from conflict.needed, which records the full demand).
    conflict_needed: dict[str, int] = {}
    for c in result.conflicts:
        conflict_needed[c.queue] = max(conflict_needed.get(c.queue, 0), c.needed)

    queue_cols: dict[str, int] = {}
    for q in queues:
        max_assigned = max(
            (len(v) for (_, _, qq), v in grid.items() if qq == q),
            default=1,
        )
        queue_cols[q] = max(1, max_assigned, conflict_needed.get(q, 0))

    total_cols_per_day = sum(queue_cols[q] for q in queues)

    # ---- Row 1: Day headers (each spans total_cols_per_day) ----
    time_col = 1
    ws.cell(row=1, column=time_col, value="Time").font = Font(bold=True)
    col = 2
    for d_idx, day in enumerate(days_present):
        colour = _DAY_COLORS[d_idx % len(_DAY_COLORS)]
        fill = PatternFill("solid", fgColor=colour)
        end_col = col + total_cols_per_day - 1
        if total_cols_per_day > 1:
            ws.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=end_col
            )
        cell = ws.cell(row=1, column=col)
        cell.value = day.upper()
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill
        col += total_cols_per_day

    # ---- Row 2: Queue headers (each queue spans queue_cols[q] columns) ----
    ws.cell(row=2, column=time_col, value="Time").font = Font(bold=True)
    col = 2
    for d_idx, _day in enumerate(days_present):
        colour = _DAY_COLORS[d_idx % len(_DAY_COLORS)]
        fill = PatternFill("solid", fgColor=colour)
        for q in queues:
            n = queue_cols[q]
            end_col = col + n - 1
            if n > 1:
                ws.merge_cells(
                    start_row=2, start_column=col, end_row=2, end_column=end_col
                )
            cell = ws.cell(row=2, column=col)
            cell.value = q
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            col += n

    # ---- Data rows: one person per cell, colored by person_id ----
    person_colors: dict[str, tuple[str, str]] = {}
    all_slots = sorted(
        {a.slot for a in result.assignments} | {c.slot for c in result.conflicts}
    )
    current_row = 3
    for slot in all_slots:
        ws.cell(row=current_row, column=time_col, value=slot_label(slot)).font = Font(bold=True)
        col = 2
        for day in days_present:
            for queue in queues:
                persons = sorted(grid.get((day, slot, queue), []))
                n = queue_cols[queue]
                for i in range(n):
                    cell = ws.cell(row=current_row, column=col)
                    if i < len(persons):
                        pid = persons[i]
                        cell.value = pid
                        if pid not in person_colors:
                            person_colors[pid] = _person_bg_fg(pid)
                        bg, fg = person_colors[pid]
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(color=fg)
                        cell.alignment = Alignment(horizontal="center")
                    col += 1
        current_row += 1

    # ---- Auto-width columns ----
    for col_cells in ws.columns:
        max_len = max([len(str(c.value or "")) for c in col_cells], default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 2, _MAX_COL_WIDTH
        )

    # ---- Conflicts sheet ----
    ws2 = wb.create_sheet("Conflicts")
    ws2.append(["Day", "Time", "Queue", "Needed", "Assigned", "Reason"])
    for i in range(1, 7):
        ws2.cell(row=1, column=i).font = Font(bold=True)
    for c in sorted(
        result.conflicts, key=lambda x: (DAYS.index(x.day), x.slot, x.queue)
    ):
        ws2.append(
            [c.day.upper(), slot_label(c.slot), c.queue, c.needed, c.assigned, c.reason]
        )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
