from __future__ import annotations

import csv
import io
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rotagen.models import DAYS, SLOTS_PER_DAY, slot_label


_DAY_COLORS = ["D6E4F7", "D6F7E4", "F7F0D6", "F7D6D6", "E4D6F7", "D6F7F3"]

# Map slot label (e.g. "08:00-08:30") -> slot index for import round-trip.
_LABEL_TO_SLOT: dict[str, int] = {slot_label(s): s for s in range(SLOTS_PER_DAY)}


def export_matrix_xlsx(demand: list[dict], queues: list[str]) -> bytes:
    """Export the requirement matrix to an XLSX workbook and return the raw bytes.

    Parameters
    ----------
    demand:
        List of ``{"day", "slot", "queue", "required"}`` dicts.
    queues:
        Ordered list of queue names used for column ordering.
    """
    lookup: dict[tuple[str, int, str], int] = {}
    for d in demand:
        lookup[(d["day"], int(d["slot"]), d["queue"])] = max(0, int(d.get("required") or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirement Matrix"

    num_queues = max(len(queues), 1)

    # Row 1: merged day headers; A1:A2 merged for "Slot"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    slot_cell = ws.cell(row=1, column=1, value="Slot")
    slot_cell.font = Font(bold=True)
    slot_cell.alignment = Alignment(horizontal="center", vertical="center")

    col = 2
    for d_idx, day in enumerate(DAYS):
        fill = PatternFill("solid", fgColor=_DAY_COLORS[d_idx % len(_DAY_COLORS)])
        if num_queues > 1:
            ws.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + num_queues - 1
            )
        cell = ws.cell(row=1, column=col)
        cell.value = day.upper()
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill
        col += num_queues

    # Row 2: queue sub-headers
    col = 2
    for d_idx, day in enumerate(DAYS):
        fill = PatternFill("solid", fgColor=_DAY_COLORS[d_idx % len(_DAY_COLORS)])
        for q in queues:
            cell = ws.cell(row=2, column=col)
            cell.value = q
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            col += 1

    # Data rows (one per slot)
    for slot in range(SLOTS_PER_DAY):
        row = slot + 3
        ws.cell(row=row, column=1, value=slot_label(slot)).font = Font(bold=True)
        col = 2
        for day in DAYS:
            for q in queues:
                ws.cell(row=row, column=col, value=lookup.get((day, slot, q), 0))
                col += 1

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_len + 2, 20
        )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def import_matrix(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse a CSV or XLSX requirement matrix and return demand rows.

    Accepted formats
    ----------------
    CSV
        Header row: ``Slot,{day}|{queue},...``  (pipe-separated day and queue in
        each column header).  Data rows follow with numeric demand values.

    XLSX
        Row 1: day headers (merged across their queue columns, e.g. ``MON``).
        Row 2: queue sub-headers per day.
        Column A: slot labels matching ``HH:MM-HH:MM`` format.
        Data rows: rows 3+.

    Returns
    -------
    list[dict]
        Each dict has keys ``day``, ``slot``, ``queue``, ``required``.
    """
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _import_xlsx(file_bytes)
    return _import_csv(file_bytes)


def _import_csv(raw: bytes) -> list[dict]:
    text = raw.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    headers = next(reader, None)
    if not headers:
        raise ValueError("Empty CSV")

    # columns[0] is None (the Slot column); subsequent entries are (day, queue) tuples.
    columns: list[tuple[str, str] | None] = [None]
    for h in headers[1:]:
        parts = h.split("|", 1)
        if len(parts) != 2:
            raise ValueError(
                f"Invalid column header {h!r}. Expected format: 'day|queue' "
                "(e.g. 'mon|Queue1')."
            )
        columns.append((parts[0].strip().lower(), parts[1].strip()))

    demand: list[dict] = []
    for row in reader:
        if not row:
            continue
        slot_lbl = row[0].strip()
        slot_idx = _LABEL_TO_SLOT.get(slot_lbl)
        if slot_idx is None:
            continue  # skip unrecognised labels and blank rows
        for col_idx, col_meta in enumerate(columns[1:], start=1):
            if col_meta is None:
                continue
            day, queue = col_meta
            try:
                required = max(0, int(float(row[col_idx])))
            except (IndexError, ValueError):
                required = 0
            demand.append({"day": day, "slot": slot_idx, "queue": queue, "required": required})
    return demand


def _import_xlsx(raw: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        raise ValueError(
            "XLSX must have at least 3 rows: day header row, queue header row, and data rows."
        )

    day_row = rows[0]
    queue_row = rows[1]

    # Build column metadata using fill-forward for merged day cells (read_only
    # mode returns None for cells that are part of a merge region).
    columns: list[tuple[str, str] | None] = [None]  # column 0 is the Slot label
    current_day: str | None = None
    for col_idx in range(1, len(queue_row)):
        day_val = day_row[col_idx] if col_idx < len(day_row) else None
        if day_val is not None:
            current_day = str(day_val).strip().lower()
        queue_val = queue_row[col_idx] if col_idx < len(queue_row) else None
        if queue_val is not None and current_day is not None:
            columns.append((current_day, str(queue_val).strip()))
        else:
            columns.append(None)

    demand: list[dict] = []
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        slot_lbl = str(row[0]).strip()
        slot_idx = _LABEL_TO_SLOT.get(slot_lbl)
        if slot_idx is None:
            continue
        for col_idx, col_meta in enumerate(columns[1:], start=1):
            if col_meta is None:
                continue
            day, queue = col_meta
            try:
                val = row[col_idx] if col_idx < len(row) else None
                required = max(0, int(float(str(val or 0))))
            except (ValueError, TypeError):
                required = 0
            demand.append({"day": day, "slot": slot_idx, "queue": queue, "required": required})
    return demand
